# Выгрузка всех заявок в файл .xlsx через openpyxl.
# Файл собирается во временном месте, отправляется админу и удаляется —
# постоянно хранить его не нужно, база всегда под рукой.

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import database

# Колонки выгрузки: заголовок и ширина в символах
COLUMNS = [
    ("№", 6),
    ("Имя", 20),
    ("Телефон", 18),
    ("Задача", 60),
    ("Telegram ID", 14),
    ("Username", 18),
    ("Создана", 20),
    ("Статус", 12),
]


def export_orders_to_xlsx(path: Path) -> int:
    """Сохранить все заявки в xlsx-файл. Возвращает число выгруженных заявок."""
    orders = database.get_all_orders()

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    ws.append([title for title, _ in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in orders:
        ws.append([
            row["id"],
            row["name"],
            row["phone"],
            row["task"],
            row["tg_user_id"],
            f"@{row['tg_username']}" if row["tg_username"] else "",
            row["created_at"],
            row["status"],
        ])

    wb.save(path)
    return len(orders)
